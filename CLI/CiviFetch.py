import os
import re
import sys
import time
import requests
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from requests.exceptions import ChunkedEncodingError, HTTPError

# ===[ 0. Setup and Utility Functions ]===
def print_banner():
    print("""
-------------------------------------------------------
 CiviFetch - CivitAI Model Downloader and Cataloger
-------------------------------------------------------
    """)

def create_session(api_key):
    """Create and configure requests session with retries"""
    session = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update({
        "User-Agent": "python-requests/2.x",
        "Authorization": f"Bearer {api_key}"
    })
    return session

def fetch_json(session, path, params=None):
    """Fetch JSON from API endpoint"""
    url = "https://civitai.com/api/v1" + path
    r = session.get(url, params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def download_with_progress(session, url, fallback, out_dir):
    """Download file with progress bar and proper retries"""
    for attempt in range(1, 6):
        try:
            resp = session.get(url, stream=True, timeout=30)
            resp.raise_for_status()
            
            # Get filename from headers or use fallback
            cd = resp.headers.get("content-disposition", "")
            m = re.search(r'filename="([^"]+)"', cd)
            fname = m.group(1) if m else fallback
            out_path = os.path.join(out_dir, fname)
            
            # Skip if file already exists
            if os.path.exists(out_path):
                print(f"  File already exists, skipping: {fname}")
                return fname
            
            # Download with progress bar
            total = int(resp.headers.get("content-length", 0))
            with open(out_path, "wb") as f, tqdm(
                total=total, unit="B", unit_scale=True,
                desc=f"Downloading {fname}"
            ) as bar:
                for chunk in resp.iter_content(8192):
                    if chunk:
                        f.write(chunk)
                        bar.update(len(chunk))
            return fname
        except (ChunkedEncodingError, HTTPError) as e:
            print(f"  [Attempt {attempt}/5] download error: {e}")
            time.sleep(2 ** (attempt-1))
    return None

def style_excel(filename):
    """Apply styling to Excel file to make it beautiful"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Style header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style data rows and apply conditional formatting
    for row in range(2, ws.max_row + 1):
        # Apply borders to all cells
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Center-align certain columns
            if col in [1, 2]:  # S.No and Model ID columns
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get status cell and apply conditional formatting
        status_cell = ws.cell(row=row, column=ws.max_column)
        if "Success" in str(status_cell.value):
            status_cell.fill = success_fill
        elif "Failed" in str(status_cell.value) or "ERROR" in str(status_cell.value):
            status_cell.fill = error_fill
        elif "Skipped" in str(status_cell.value):
            status_cell.fill = skip_fill
    
    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        # Set custom widths based on column content
        if col == 1:  # S.No
            ws.column_dimensions[column_letter].width = 8
        elif col == 2:  # Model ID
            ws.column_dimensions[column_letter].width = 12
        elif col == 3:  # Model Name
            ws.column_dimensions[column_letter].width = 40
        elif col == 4:  # Tags
            ws.column_dimensions[column_letter].width = 30
        elif col == 5:  # Trigger words
            ws.column_dimensions[column_letter].width = 50
        elif col == 6:  # Base Model
            ws.column_dimensions[column_letter].width = 15
        elif col in [7, 8]:  # Hash columns
            ws.column_dimensions[column_letter].width = 20
        elif col == 9:  # File Size
            ws.column_dimensions[column_letter].width = 15
        elif col == 10:  # NSFW
            ws.column_dimensions[column_letter].width = 8
        else:  # Status
            ws.column_dimensions[column_letter].width = 30
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary at the top
    ws.insert_rows(1, 2)
    ws.merge_cells('A1:K1')
    summary_cell = ws.cell(row=1, column=1)
    total_models = ws.max_row - 3  # Accounting for header and this summary row
    
    # Count statuses
    success_count = 0
    failed_count = 0
    skipped_count = 0
    
    for row in range(4, ws.max_row + 1):
        status = ws.cell(row=row, column=ws.max_column).value
        if status and "Success" in status:
            success_count += 1
        elif status and ("Failed" in status or "ERROR" in status):
            failed_count += 1
        elif status and "Skipped" in status:
            skipped_count += 1
    
    summary_cell.value = f"CivitAI Models Summary | Total: {total_models} | Success: {success_count} | Failed: {failed_count} | Skipped: {skipped_count} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_cell.font = Font(name="Calibri", size=12, bold=True)
    summary_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_cell.border = border
    
    # Save the styled workbook
    wb.save(filename)
    return filename

def detect_nsfw_from_tags(tags, nsfw_flag=False):
    """Detect if a model is NSFW based on tags or default flag"""
    if not tags:
        return "SFW"
    
    # Check if any tag contains "nsfw" (case insensitive)
    for tag in tags:
        if isinstance(tag, str) and "nsfw" in tag.lower():
            return "NSFW"
    
    # If model is explicitly flagged as NSFW but no NSFW tags
    if nsfw_flag:
        return "NSFW"
        
    return "SFW"

# ===[ 1. Main Program ]===
def main():
    print_banner()
    
    # Prompt for API Key & NSFW Warning
    API_KEY = input("Enter your CivitAI API key: ").strip()
    if not API_KEY:
        print("Error: API key is required.")
        sys.exit(1)

    print("\nMake sure your CivitAI account has NSFW enabled.")
    print("NSFW models will be skipped if your key cannot access them.\n")

    # Choose download mode
    print("Choose download mode:")
    print("  1) Single model (Paste one URL or ID)")
    print("  2) Bulk models (Read links/IDs from a text file, One Url per line)")
    print()
    print("NOTE: SFW OR NSFW TAGS ARE GIVEN IN THE EXCEL FILE BASED ON THERE PRESENCE AT THE MODEL-TAGS. IF IN MODEL-TAG NSFW IS PRESENT IT WILL MARK MODEL AS NSFW IN EXCEL FILE. IF SFW PRESENT IN MODEL-TAG THEN WILL MARK SFW IN EXCEL FILE OR NO TAGS PRESENT IT WILL DEFAULT TAG AS SFW IN EXCEL FILE. SO RECECK YOUR MODELS FOR THERE NSFW/SFW PROPERTY")
    print()
    print("WARNING: THIS IS AN OPEN-SOURCE PROJECT TO HELP YOU WITH MODEL DOWNLOADS, SINGLE OR BULK MANAGEMENT, AND SAVING MODEL DETAILS. WHAT YOU DO WITH IT IS ENTIRELY YOUR RESPONSIBILITY. I AM NOT LIABLE FOR YOUR ACTIONS. I HAVE ONLY PROVIDED THE TOOL — USE IT RESPONSIBLY AND LEGALLY.")
    print()
    mode = input("Enter 1 or 2: ").strip()

    entries = []
    if mode == "1":
        val = input("Paste the model URL or numeric ID: ").strip()
        if val:
            entries = [val]
    elif mode == "2":
        path = input("Enter path to TXT file (one URL/ID per line): ").strip()
        if not os.path.isfile(path):
            print("File not found:", path)
            sys.exit(1)
        with open(path, "r", encoding="utf-8") as f:
            entries = [line.strip() for line in f if line.strip()]
    else:
        print("Invalid choice.")
        sys.exit(1)

    if not entries:
        print("No models to process.")
        sys.exit(0)

    # Remove duplicates and extract model IDs
    unique_model_ids = []
    seen_ids = set()
    
    for entry in entries:
        # Extract numeric model ID from URL or accept ID
        m = re.search(r"/models/(\d+)", entry)
        model_id = m.group(1) if m else entry
        
        if model_id not in seen_ids:
            seen_ids.add(model_id)
            unique_model_ids.append(model_id)

    print(f"Found {len(unique_model_ids)} unique models out of {len(entries)} entries.")

    # Prepare output folders
    os.makedirs("CivitModels", exist_ok=True)
    os.makedirs("CivitData", exist_ok=True)

    # Create HTTP session
    session = create_session(API_KEY)
    
    # ===[ Phase 1: Collect all metadata first ]===
    print("\n------- PHASE 1: COLLECTING METADATA -------")
    all_metadata = []  # List to hold metadata for all models
    
    for idx, model_id in enumerate(unique_model_ids, 1):
        print(f"\nFetching metadata {idx}/{len(unique_model_ids)}: ID {model_id}")
        
        try:
            md = fetch_json(session, f"/models/{model_id}")
            
            # Extract version and file info
            ver = md["modelVersions"][0]
            file0 = ver["files"][0]
            hashes = file0.get("hashes", {})
            
            # Extract basic fields
            model_name = md.get("name", "—")
            trigger_words = ver.get("trainedWords", [])
            base_model = ver.get("baseModel", "—")
            sha256_hash = hashes.get("SHA256", "—")
            autov1_hash = hashes.get("AutoV1", "—")
            file_size = file0.get("sizeKB", 0)
            file_size_formatted = f"{file_size/1024:.2f} MB" if file_size else "—"
            
            # Extract tags and check for NSFW
            tags = md.get("tags", [])
            is_nsfw_flagged = md.get("nsfw", False)
            nsfw_status = detect_nsfw_from_tags(tags, is_nsfw_flagged)
            
            # Store download URL for later
            dl_url = file0["downloadUrl"] + f"?token={API_KEY}"
            
            all_metadata.append({
                "S.No": idx,
                "Model ID": model_id,
                "Model Name": model_name,
                "Tags": ", ".join(tags) if tags else "—",
                "Trigger Words": "; ".join(trigger_words) or "—",
                "Base Model": base_model,
                "SHA256": sha256_hash,
                "AutoV1": autov1_hash,
                "File Size": file_size_formatted,
                "NSFW": nsfw_status,
                "Status": "Pending",
                # Hidden metadata for download phase
                "_download_url": dl_url,
                "_filename": f"{model_name.replace(' ','_')}.safetensors",
                "_is_nsfw": is_nsfw_flagged or nsfw_status == "NSFW"
            })
            
            print(f"  Got metadata for '{model_name}' ({base_model}, {file_size_formatted})")
            
        except HTTPError as e:
            print(f"  Failed to fetch metadata: {e}")
            all_metadata.append({
                "S.No": idx,
                "Model ID": model_id,
                "Model Name": "ERROR - Failed to fetch",
                "Tags": "—",
                "Trigger Words": "—",
                "Base Model": "—",
                "SHA256": "—",
                "AutoV1": "—",
                "File Size": "—",
                "NSFW": "—",
                "Status": "Failed to fetch metadata",
                # No download info
                "_download_url": None,
                "_filename": None,
                "_is_nsfw": False
            })
            continue
    
    # Generate first Excel file with just metadata
    print("\nCreating initial metadata Excel file...")
    
    # Remove internal fields for Excel display
    display_metadata = []
    for item in all_metadata:
        display_item = item.copy()
        for k in list(display_item.keys()):
            if k.startswith("_"):
                del display_item[k]
        display_metadata.append(display_item)
    
    df = pd.DataFrame(display_metadata)
    
    # Name the Excel file based on number of links
    excel_filename = f"CivitData/{len(unique_model_ids)}_models.xlsx"
    df.to_excel(excel_filename, index=False, engine='openpyxl')
    
    # Style the Excel file to make it beautiful
    styled_file = style_excel(excel_filename)
    print(f"  Created initial metadata Excel file: {styled_file}")
    
    # ===[ Phase 2: Download models ]===
    print("\n------- PHASE 2: DOWNLOADING MODELS -------")
    
    download_choice = input("\nStart downloading models? (Y/N): ").strip().lower()
    if download_choice == "n":
        print("Downloads skipped. Process complete.")
        return
    
    # Single prompt for NSFW handling
    nsfw_models = [model for model in all_metadata if model.get("_is_nsfw", False)]
    if nsfw_models:
        print(f"\nFound {len(nsfw_models)} NSFW models in your list.")
        nsfw_choice = input("Download NSFW models? (Y/N): ").strip().lower()
        download_nsfw = nsfw_choice == "y"
    else:
        download_nsfw = False
    
    for idx, model_data in enumerate(all_metadata, 1):
        model_id = model_data["Model ID"]
        model_name = model_data["Model Name"]
        dl_url = model_data.get("_download_url")
        filename = model_data.get("_filename")
        is_nsfw = model_data.get("_is_nsfw", False)
        
        if not dl_url:
            print(f"\nSkipping model {idx}/{len(all_metadata)}: ID {model_id} (no download URL)")
            continue
        
        print(f"\nProcessing download {idx}/{len(all_metadata)}: {model_name} (ID: {model_id})")
        
        # Handle NSFW content based on global choice
        if is_nsfw and not download_nsfw:
            print(f"  Model is NSFW. Skipping based on your preference.")
            model_data["Status"] = "Skipped (NSFW)"
            continue
        
        # Download file
        downloaded = download_with_progress(session, dl_url, filename, "CivitModels")
        if downloaded:
            print(f"  Model saved to CivitModels/{downloaded}")
            model_data["Status"] = f"Success - {downloaded}"
        else:
            print(f"  Download failed for model {model_id}")
            model_data["Status"] = "Download failed"
    
    # Generate final Excel file with updated statuses
    print("\nCreating final metadata Excel file with download statuses...")
    
    # Remove internal fields for Excel display again
    final_display_metadata = []
    for item in all_metadata:
        display_item = item.copy()
        for k in list(display_item.keys()):
            if k.startswith("_"):
                del display_item[k]
        final_display_metadata.append(display_item)
    
    df = pd.DataFrame(final_display_metadata)
    df.to_excel(excel_filename, index=False, engine='openpyxl')
    
    # Style the Excel file to make it beautiful
    styled_file = style_excel(excel_filename)
    print(f"  Created final metadata Excel file: {styled_file}")
    
    print("\nAll tasks complete!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProcess interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)