import os
import io
import sys 
import re
import time
import requests
import pandas as pd
import json
import gradio as gr

from datetime import datetime
from io import BytesIO
from PIL import Image
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from requests.exceptions import RequestException, Timeout, ConnectionError
from typing import Dict, Tuple, Optional, Any, List


# Custom Hugging Face theme
hf_theme = gr.themes.Default(
    primary_hue="orange",
    secondary_hue="blue",
    font=[gr.themes.GoogleFont("Inter")],
).set(
    body_background_fill='#0b0f19',
    button_primary_background_fill='#ff8b3d',
    button_primary_background_fill_hover='#ff6b3d',
    block_background_fill='#1a1f2d',
    input_background_fill='#252a38',
    border_color_primary='#353945',
)

# ===[ Utility Functions ]===
def create_session(api_key):
    """Create and configure requests session with retries"""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update({
        "User-Agent": "python-requests/2.x",
        "Authorization": f"Bearer {api_key}"
    })
    return session

def fetch_json(session, path, params=None, timeout=30):
    """Fetch JSON from API endpoint with increased timeout"""
    url = "https://civitai.com/api/v1" + path
    r = session.get(url, params=params, timeout=timeout)
    r.raise_for_status()
    return r.json()

def format_size(bytes_size):
    """Format file size to appropriate units"""
    if bytes_size < 1024:
        return f"{bytes_size} B"
    elif bytes_size < 1024**2:
        return f"{bytes_size/1024:.1f} KB"
    elif bytes_size < 1024**3:
        return f"{bytes_size/1024**2:.1f} MB"
    else:
        return f"{bytes_size/1024**3:.2f} GB"

def download_with_progress(session, url, fallback, out_dir, progress=None):
    """Download file with progress tracking and increased timeout"""
    for attempt in range(1, 4):
        try:
            resp = session.get(url, stream=True, timeout=60)
            resp.raise_for_status()
            
            # Get filename from headers or use fallback
            cd = resp.headers.get("content-disposition", "")
            m = re.search(r'filename="([^"]+)"', cd)
            fname = m.group(1) if m else fallback
            out_path = os.path.join(out_dir, fname)
            
            # Skip if file already exists
            if os.path.exists(out_path):
                return f"File already exists, skipping: {fname}"
            
            # Download with progress tracking
            total = int(resp.headers.get("content-length", 0))
            
            with open(out_path, "wb") as f:
                downloaded = 0
                last_update = 0
                progress_update_threshold = int(total * 0.25) if total > 0 else 0
                
                for chunk in resp.iter_content(8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        
                        if total > 0 and (downloaded - last_update > progress_update_threshold or downloaded == total):
                            last_update = downloaded
                            progress_pct = int(100 * downloaded / total)
                            progress(f"Downloading {fname}: {progress_pct}% ({format_size(downloaded)} / {format_size(total)})")
            
            return f"Successfully downloaded: {fname}"
            
        except RequestException as e:
            error_msg = f"[Attempt {attempt}/3] download error: {e}"
            if progress:
                progress(error_msg)
            time.sleep(2 ** (attempt-1))
    
    return "Download failed after multiple attempts"

def style_excel(filename):
    """Apply styling to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Style header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Center-align certain columns
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        status_cell = ws.cell(row=row, column=ws.max_column)
        if "Success" in str(status_cell.value):
            status_cell.fill = success_fill
        elif "Failed" in str(status_cell.value) or "ERROR" in str(status_cell.value):
            status_cell.fill = error_fill
        elif "Skipped" in str(status_cell.value):
            status_cell.fill = skip_fill
    
    # Adjust column widths
    column_widths = [8, 12, 40, 30, 50, 15, 20, 20, 15, 8, 30]
    for col, width in enumerate(column_widths[:ws.max_column], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary at the top
    ws.insert_rows(1, 2)
    ws.merge_cells('A1:K1')
    summary_cell = ws.cell(row=1, column=1)
    
    # Count statuses
    total_models = ws.max_row - 3
    success_count = sum(1 for row in range(4, ws.max_row + 1) 
                      if "Success" in str(ws.cell(row=row, column=ws.max_column).value or ""))
    failed_count = sum(1 for row in range(4, ws.max_row + 1) 
                     if any(x in str(ws.cell(row=row, column=ws.max_column).value or "") 
                            for x in ["Failed", "ERROR"]))
    skipped_count = sum(1 for row in range(4, ws.max_row + 1) 
                      if "Skipped" in str(ws.cell(row=row, column=ws.max_column).value or ""))
    
    summary_cell.value = f"CivitAI Models Summary | Total: {total_models} | Success: {success_count} | Failed: {failed_count} | Skipped: {skipped_count} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_cell.font = Font(name="Calibri", size=12, bold=True)
    summary_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_cell.border = border
    
    # Save the styled workbook
    wb.save(filename)
    return filename

def detect_nsfw_from_tags(tags, nsfw_flag=False):
    """Detect if a model is NSFW based on tags or default flag"""
    if nsfw_flag:
        return "NSFW"
        
    if not tags:
        return "SFW"
    
    # Check if any tag contains "nsfw" (case insensitive)
    for tag in tags:
        if isinstance(tag, str) and "nsfw" in tag.lower():
            return "NSFW"
    
    return "SFW"

def sanitize_filename(name):
    """Sanitize string for use as filename."""
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
    return sanitized[:100]

def parse_url(url):
    """Parse model ID and version ID from URL."""
    model_match = re.search(r"/models/(\d+)", url)
    version_match = re.search(r"modelVersionId=(\d+)", url)
    
    if not model_match:
        raise ValueError("Error: Invalid model URL format. Expected URL like https://civitai.com/models/XXXXX")
        
    model_id = model_match.group(1)
    version_id = version_match.group(1) if version_match else None
    
    return model_id, version_id

# ===[ Model Download Functions ]===
def handle_model_download_new(api_key, mode, model_input, file_input, nsfw_choice):
    """Process model download request"""
    # Translate nsfw_choice to Yes/No for compatibility with existing function
    nsfw_toggle = "Yes" if nsfw_choice == "NSFW Included" else "No"
    
    # Create required directories
    os.makedirs("CivitModels", exist_ok=True)
    os.makedirs("CivitData", exist_ok=True)
    
    # Create session
    session = create_session(api_key)
    
    progress_text = ""
    
    def progress_callback(text):
        nonlocal progress_text
        progress_text += text + "\n"
    
    # Extract model IDs based on mode
    unique_model_ids = []
    
    if mode == "single":
        try:
            progress_callback("Parsing model URL...")
            # Extract numeric model ID from URL or accept ID
            m = re.search(r"/models/(\d+)", model_input)
            model_id = m.group(1) if m else model_input.strip()
            unique_model_ids.append(model_id)
        except Exception as e:
            return f"Error parsing model URL: {str(e)}"
    else:  # bulk mode
        try:
            progress_callback("Reading model URLs from file...")
            with open(file_input, "r", encoding="utf-8") as f:
                entries = [line.strip() for line in f if line.strip()]
                
            if not entries:
                return "No models found in the uploaded file."
                
            seen_ids = set()
            for entry in entries:
                m = re.search(r"/models/(\d+)", entry)
                model_id = m.group(1) if m else entry
                
                if model_id not in seen_ids:
                    seen_ids.add(model_id)
                    unique_model_ids.append(model_id)
                    
            progress_callback(f"Found {len(unique_model_ids)} unique models out of {len(entries)} entries.")
        except Exception as e:
            return f"Error processing file: {str(e)}"
    
    # Process models
    download_nsfw = nsfw_toggle == "Yes"
    all_metadata = []
    valid_models = []  # Track models that pass NSFW filter
    nsfw_models = []   # Track NSFW models specifically
    
    # Phase 1: Collect metadata
    progress_callback("\n===== PHASE 1: COLLECTING METADATA =====")
    
    for idx, model_id in enumerate(unique_model_ids, 1):
        progress_callback(f"\nFetching metadata {idx}/{len(unique_model_ids)}: ID {model_id}")
        
        try:
            md = fetch_json(session, f"/models/{model_id}")
            
            # Extract version and file info
            ver = md["modelVersions"][0]
            file0 = ver["files"][0]
            hashes = file0.get("hashes", {})
            
            # Extract basic fields
            model_name = md.get("name", "—")
            trigger_words = ver.get("trainedWords", [])
            base_model = ver.get("baseModel", "—")
            sha256_hash = hashes.get("SHA256", "—")
            autov1_hash = hashes.get("AutoV1", "—")
            file_size = file0.get("sizeKB", 0)
            file_size_formatted = f"{file_size/1024:.2f} MB" if file_size else "—"
            
            # Extract tags and check for NSFW
            tags = md.get("tags", [])
            is_nsfw_flagged = md.get("nsfw", False)
            nsfw_status = detect_nsfw_from_tags(tags, is_nsfw_flagged)
            
            model_data = {
                "S.No": idx,
                "Model ID": model_id,
                "Model Name": model_name,
                "Tags": ", ".join(tags) if tags else "—",
                "Trigger Words": "; ".join(trigger_words) or "—",
                "Base Model": base_model,
                "SHA256": sha256_hash,
                "AutoV1": autov1_hash,
                "File Size": file_size_formatted,
                "NSFW": nsfw_status,
                "Status": "Pending",
                # Hidden metadata for download phase
                "_download_url": file0["downloadUrl"] + f"?token={api_key}",
                "_filename": sanitize_filename(f"{model_name}.safetensors"),
                "_is_nsfw": is_nsfw_flagged or nsfw_status == "NSFW"
            }
            
            # Check NSFW policy - separate NSFW models
            if nsfw_status == "NSFW":
                if download_nsfw:
                    progress_callback(f"  Model '{model_name}' is NSFW. Will download based on your preference.")
                    valid_models.append(model_data)
                    nsfw_models.append(model_data)
                else:
                    progress_callback(f"  Model '{model_name}' is NSFW. Skipping based on your preference.")
                    model_data["Status"] = "Skipped (NSFW)"
                    nsfw_models.append(model_data)  # Add to NSFW list, not all_metadata
            else:
                # SFW model - always add to valid and all list
                progress_callback(f"  Got metadata for '{model_name}' ({base_model}, {file_size_formatted})")
                valid_models.append(model_data)
                all_metadata.append(model_data)
            
        except Exception as e:
            progress_callback(f"  Failed to fetch metadata: {e}")
            error_model = {
                "S.No": idx,
                "Model ID": model_id,
                "Model Name": "ERROR - Failed to fetch",
                "Tags": "—",
                "Trigger Words": "—",
                "Base Model": "—",
                "SHA256": "—",
                "AutoV1": "—",
                "File Size": "—",
                "NSFW": "—",
                "Status": f"Failed: {str(e)}",
            }
            all_metadata.append(error_model)
            continue
    
    # Add NSFW models to all_metadata only if downloading NSFW
    if download_nsfw and nsfw_models:
        all_metadata.extend(nsfw_models)
    
    # Early exit if no valid models to download
    if not valid_models:
        progress_callback("\nNo valid models to download after applying NSFW filter.")
        
        # Only create Excel if we have models to report
        if all_metadata:
            display_metadata = []
            for item in all_metadata:
                display_item = {k: v for k, v in item.items() if not k.startswith("_")}
                display_metadata.append(display_item)
                
            df = pd.DataFrame(display_metadata)
            excel_filename = f"CivitData/{len(all_metadata)}_models_report.xlsx"
            df.to_excel(excel_filename, index=False, engine='openpyxl')
            styled_file = style_excel(excel_filename)
            progress_callback(f"Created report for models: {styled_file}")
        
        return progress_text
    
    # Generate initial Excel file
    progress_callback("\nCreating initial metadata Excel file...")
    
    # Remove internal fields for Excel display
    display_metadata = []
    for item in all_metadata:
        display_item = {k: v for k, v in item.items() if not k.startswith("_")}
        display_metadata.append(display_item)
    
    df = pd.DataFrame(display_metadata)
    excel_filename = f"CivitData/{len(valid_models)}_models.xlsx"
    df.to_excel(excel_filename, index=False, engine='openpyxl')
    styled_file = style_excel(excel_filename)
    progress_callback(f"  Created initial metadata Excel file: {styled_file}")
    
    # Phase 2: Download models
    progress_callback("\n===== PHASE 2: DOWNLOADING MODELS =====")
    
    for idx, model_data in enumerate(valid_models, 1):
        model_id = model_data["Model ID"]
        model_name = model_data["Model Name"]
        dl_url = model_data.get("_download_url")
        filename = model_data.get("_filename")
        
        progress_callback(f"\nProcessing download {idx}/{len(valid_models)}: {model_name} (ID: {model_id})")
        
        # Download file
        result = download_with_progress(session, dl_url, filename, "CivitModels", progress_callback)
        progress_callback(result)
        
        if "Successfully" in result:
            model_data["Status"] = f"Success - {filename}"
        elif "already exists" in result:
            model_data["Status"] = f"Skipped (Already exists)"
        else:
            model_data["Status"] = f"Failed: {result}"
    
    # Generate final Excel file
    progress_callback("\nCreating final metadata Excel file with download statuses...")
    
    # Remove internal fields for Excel display again
    final_display_metadata = []
    for item in all_metadata:
        display_item = {k: v for k, v in item.items() if not k.startswith("_")}
        final_display_metadata.append(display_item)
    
    df = pd.DataFrame(final_display_metadata)
    df.to_excel(excel_filename, index=False, engine='openpyxl')
    styled_file = style_excel(excel_filename)
    progress_callback(f"  Created final metadata Excel file: {styled_file}")
    
    final_message = f"All tasks complete! Downloaded models are in 'CivitModels' folder. Report available at {styled_file}"
    progress_callback(final_message)
    
    return progress_text

# ===[ Image Download Function ]===    
ROOT_FOLDER = "CivitImg"
API_BASE = "https://civitai.com/api/v1"
MAX_RETRIES = 3
RETRY_DELAY = 10
REQUEST_TIMEOUT = 30

# ===[ Print Capture System ]===
# Ensure this class is defined before it's used
class PrintCapture:
    def __init__(self):
        self.buffer = []

    def write(self, text):
        # Capture standard output and errors
        self.buffer.append(str(text))

    def flush(self):
        # Required for file-like object interface
        pass

    def get_output(self):
        return "".join(self.buffer)

    def clear(self):
        self.buffer = []

# Instantiate PrintCapture globally or where accessible by handle_image_download
output_capture = PrintCapture()

# ===[ Image Download Class ]===
class CivitaiDownloader:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.headers = {'Authorization': f'Bearer {api_key}'}
        # Use the PrintCapture instance for logging within the class
        self.logger = output_capture
        os.makedirs(ROOT_FOLDER, exist_ok=True)

    def parse_url(self, url: str) -> Tuple[str, Optional[str]]:
        """Parse model ID and version ID from URL."""
        model_match = re.search(r"/models/(\d+)", url)
        version_match = re.search(r"modelVersionId=(\d+)", url)

        if not model_match:
            raise ValueError("Invalid model URL format. Expected like https://civitai.com/models/XXXXX")

        model_id = model_match.group(1)
        version_id = version_match.group(1) if version_match else None

        return model_id, version_id

    def fetch_model_info(self, model_id: str, version_id: Optional[str] = None) -> Dict:
        """Fetch model information including name."""
        try:
            model_data = self.make_api_request(f"models/{model_id}")
            model_name = model_data.get("name", f"model_{model_id}")
            version_name = None # Default

            if version_id:
                # Find specific version name if version_id is provided
                for version in model_data.get("modelVersions", []):
                    if str(version.get("id")) == version_id:
                        version_name = version.get("name", f"version_{version_id}")
                        break
                if not version_name: # Fallback if specific version not found by ID somehow
                     version_name = f"version_{version_id}"

            return {
                "model_name": model_name,
                "version_name": version_name # Can be None if no version_id was passed
            }
        except Exception as e:
            self.logger.write(f"Error fetching model info: {str(e)}\n")
            # Provide default names even on error to allow download attempt
            return {
                "model_name": f"model_{model_id}",
                "version_name": f"version_{version_id}" if version_id else None
            }

    def sanitize_filename(self, name: str) -> str:
        """Sanitize string for use as filename."""
        # Remove invalid characters
        sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
        # Remove leading/trailing whitespace and dots
        sanitized = sanitized.strip('. ')
        # Limit length
        return sanitized[:150] # Increased length slightly for potentially long names

    def make_api_request(self, endpoint: str, params: Dict[str, Any] = None) -> Optional[Dict]:
        """Make API request with retry logic."""
        url = f"{API_BASE}/{endpoint}"
        retries = 0

        while retries < MAX_RETRIES:
            try:
                response = requests.get(
                    url,
                    headers=self.headers,
                    params=params,
                    timeout=REQUEST_TIMEOUT
                )

                if response.status_code == 200:
                    return response.json()
                elif response.status_code in [401, 403]:
                     self.logger.write(f"API Authentication Error: {response.status_code}. Check your API key.\n")
                     return None # Authentication errors are fatal
                elif response.status_code in [429, 524]: # Rate limit or Cloudflare timeout
                    wait_time = int(response.headers.get('Retry-After', RETRY_DELAY))
                    self.logger.write(f"Rate limit or timeout ({response.status_code}). Waiting {wait_time}s (Attempt {retries + 1}/{MAX_RETRIES})\n")
                    time.sleep(wait_time)
                    retries += 1
                elif response.status_code >= 500: # Server errors
                    self.logger.write(f"Server error ({response.status_code}). Retrying in {RETRY_DELAY}s (Attempt {retries + 1}/{MAX_RETRIES})\n")
                    retries += 1
                    time.sleep(RETRY_DELAY * (retries)) # Exponential backoff might be better
                else: # Other client errors (404 Not Found, etc.)
                    self.logger.write(f"API Client Error: {response.status_code} for URL {response.url}. Params: {params}\nResponse: {response.text[:200]}\n")
                    return None # Assume non-retryable client errors are fatal

            except (ConnectionError, Timeout) as e:
                self.logger.write(f"Connection error: {str(e)}. Retrying in {RETRY_DELAY}s (Attempt {retries + 1}/{MAX_RETRIES})\n")
                retries += 1
                time.sleep(RETRY_DELAY * (retries)) # Exponential backoff

        self.logger.write(f"Max retries exceeded for endpoint {endpoint}.\n")
        return None # Return None after exceeding retries


    def download_gallery(self, model_id: str, version_id: Optional[str], model_name: str, filters: Dict[str, str]):
        """Download all images from model's gallery using pagination until no more items are returned."""
        params = {
            'modelId': model_id,
            'limit': 100,  # Using 100 as a safer limit often seen in APIs
            'page': 1,
            **filters
        }

        if version_id:
            params['modelVersionId'] = version_id

        # Create a folder name combining model name and ID for uniqueness
        safe_model_name = self.sanitize_filename(model_name)
        model_folder_name = f"{safe_model_name}_(ID_{model_id})"
        if version_id:
             # Add version ID if specific version is targeted
             model_folder_name += f"_VerID_{version_id}"

        model_folder = os.path.join(ROOT_FOLDER, model_folder_name)
        os.makedirs(model_folder, exist_ok=True)
        self.logger.write(f"Saving images to folder: {model_folder}\n")

        total_downloaded = 0
        processed_image_ids = set() # Keep track of downloaded image IDs to prevent duplicates across pages

        while True: # Loop indefinitely until explicitly broken
            self.logger.write(f"Fetching gallery page {params['page']}...\n")
            images_data = self.make_api_request('images', params)

            # Handle API request failure gracefully
            if images_data is None:
                 self.logger.write(f"Failed to fetch image data for page {params['page']}. Stopping download for this model.\n")
                 break

            images = images_data.get('items', [])
            if not images:
                # This is the key condition: stop if the API returns an empty list for the current page
                self.logger.write("No more images found on this page or subsequent pages.\n")
                break

            page_download_count = 0
            for img in images:
                img_id = img.get('id')
                img_url = img.get('url')

                # Skip if no URL or if image ID has already been processed (handles potential API pagination overlap)
                if not img_url or not img_id:
                    self.logger.write(f"Skipping image data with missing URL or ID: {img}\n")
                    continue

                if img_id in processed_image_ids:
                     self.logger.write(f"Image ID {img_id} already processed, skipping (potential duplicate).\n")
                     continue

                # Determine filename (ensure unique name using image ID)
                # Try to get extension from URL, default to jpg
                url_parts = img_url.split('.')
                img_extension = url_parts[-1].lower() if len(url_parts) > 1 and len(url_parts[-1]) <= 4 else 'jpg'
                # Sanitize potentially included parameters in URL extension part
                img_extension = re.sub(r'[?#].*$', '', img_extension)
                if img_extension not in ['jpg', 'jpeg', 'png', 'webp', 'gif']: # Basic check for valid extensions
                    self.logger.write(f"Warning: Unusual extension '{img_extension}' for image ID {img_id}. Defaulting to '.jpg'. URL: {img_url}\n")
                    img_extension = 'jpg'

                img_name = f"img_{img_id}.{img_extension}"
                # Sanitize the final generated filename
                safe_img_name = self.sanitize_filename(img_name)
                img_path = os.path.join(model_folder, safe_img_name)

                if os.path.exists(img_path):
                    # self.logger.write(f"Image {safe_img_name} already exists, skipping.\n") # Less verbose logging
                    processed_image_ids.add(img_id) # Still mark as processed
                    continue

                # Download the image
                # self.logger.write(f"Downloading {safe_img_name}...\n") # Less verbose logging
                try:
                    img_response = requests.get(img_url, timeout=REQUEST_TIMEOUT, stream=True) # Use stream=True for potentially large images
                    img_response.raise_for_status() # Raise HTTPError for bad responses (4xx or 5xx)

                    with open(img_path, 'wb') as f:
                        for chunk in img_response.iter_content(chunk_size=8192):
                            f.write(chunk)

                    total_downloaded += 1
                    page_download_count += 1
                    processed_image_ids.add(img_id) # Mark as processed only after successful download

                except requests.exceptions.RequestException as e:
                    self.logger.write(f"Error downloading image ID {img_id} from {img_url}: {str(e)}\n")
                    # Optionally: remove partially downloaded file if error occurs
                    if os.path.exists(img_path):
                        try:
                            os.remove(img_path)
                        except OSError as rm_err:
                            self.logger.write(f"Could not remove partial file {img_path}: {rm_err}\n")

            self.logger.write(f"Downloaded {page_download_count} images from page {params['page']}.\n")

            # Prepare for the next page
            metadata = images_data.get('metadata', {})
            current_page = metadata.get('currentPage', params['page'])
            # Check if there's a next page hinted by the API, though primary check is empty 'items'
            next_page_url = metadata.get('nextPage')
            if not next_page_url:
                 self.logger.write("API metadata indicates no next page URL.\n")
                 # break # Removed break here, rely on empty items list as the primary signal

            params['page'] = current_page + 1 # Increment page number for the next request
            time.sleep(0.5) # Shorter delay between pages, increase if rate limited

        self.logger.write(f"Image download process complete for model '{model_name}'. Total images downloaded: {total_downloaded}\n")


# Define the missing handler function for Gradio
# === (Previous code like PrintCapture, CivitaiDownloader class remains the same) ===

# Updated handle_image_download function for "All", "SFW Only", "NSFW Only"
def handle_image_download(api_key, url, nsfw_choice):
    """Handles the image download request from the Gradio interface."""
    if not api_key:
        return "Error: API Key is required."
    if not url:
        return "Error: Model Gallery URL is required."

    output_capture.clear()
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    sys.stdout = output_capture # Redirect stdout
    sys.stderr = output_capture # Redirect stderr

    try:
        downloader = CivitaiDownloader(api_key) # Instantiates the downloader

        print(f"Parsing URL: {url}") # Use print, it's redirected
        model_id, version_id = downloader.parse_url(url)
        print(f"Parsed Model ID: {model_id}, Version ID: {version_id if version_id else 'Latest'}")

        print("Fetching model information...")
        model_info = downloader.fetch_model_info(model_id, version_id)
        model_name = model_info['model_name']
        print(f"Found Model Name: {model_name}")
        if version_id and model_info.get('version_name'):
            print(f"Targeting Version Name: {model_info['version_name']}")

        # --- Corrected NSFW Filter Logic ---
        nsfw_filter = {} # Start with empty filters
        print(f"Received NSFW choice from UI: {nsfw_choice}")

        if nsfw_choice == "NSFW Only":
            # Parameter to request ONLY NSFW images. 'true' is common.
            nsfw_filter = {'nsfw': 'true'}
            print("API Filter set to: NSFW Only ({'nsfw': 'true'})")
            # Note: If 'true' doesn't work, Civitai might use 'Only'. Check API docs if needed.
        elif nsfw_choice == "SFW Only":
            # Parameter to request ONLY SFW images (exclude NSFW). 'false' is standard.
            nsfw_filter = {'nsfw': 'false'}
            print("API Filter set to: SFW Only ({'nsfw': 'false'})")
        elif nsfw_choice == "All":
            # To get ALL images, typically you omit the nsfw parameter entirely.
            nsfw_filter = {} # Keep the filter dictionary empty
            print("API Filter set to: All (No 'nsfw' parameter sent)")

        print(f"\nStarting image download for model: {model_name} (ID: {model_id})")
        downloader.download_gallery(
            model_id=model_id,
            version_id=version_id,
            model_name=model_name,
            filters=nsfw_filter # Pass the determined filter ({} for All)
        )
        print("\nImage download function finished.")

    except ValueError as ve:
        print(f"Configuration Error: {str(ve)}")
    except Exception as e:
        import traceback
        print("\n--- UNEXPECTED ERROR ---")
        print(f"An unexpected error occurred: {str(e)}")
        print("Traceback:")
        print(traceback.format_exc())
        print("--- END TRACEBACK ---")
    finally:
        sys.stdout = original_stdout
        sys.stderr = original_stderr

    return output_capture.get_output()

# Main Gradio UI
with gr.Blocks(theme=hf_theme, title="CivitAI Fetch (Developed By Voiid)") as app:
    # Title and warning section
    gr.Markdown("# 🚀 CivitAI Fetch")
    gr.Markdown('<p style="color:#ff3860; font-weight:bold; background-color:#301b22; padding:10px; border-radius:5px; border:1px solid #ff3860">⚠️ WARNING: To download NSFW content, your CivitAI account settings must have NSFW enabled</p>', elem_id="warning")
    gr.Markdown("Developed by Voiid for Personal Use")
    
    with gr.Tabs():
        with gr.Tab("📦 Model Download"):
            with gr.Row():
                with gr.Column():
                    gr.Markdown("### Model Download Settings")
                    model_api_key = gr.Textbox(label="API Key", type="password")
                    mode = gr.Radio(["single", "bulk"], label="Download Mode", value="single")
                    
                    with gr.Group(visible=True) as single_mode_group:
                        model_input = gr.Textbox(
                            label="Model URL/ID", 
                            placeholder="https://civitai.com/models/... or Model ID"
                        )
                    
                    with gr.Group(visible=False) as bulk_mode_group:
                        gr.Markdown("📄 **Upload a text file with one URL per line:**")
                        file_input = gr.File(
                            label="Upload TXT File with URLs", 
                            file_types=[".txt"]
                        )
                    
                    nsfw_toggle = gr.Radio(
                        ["SFW Only", "NSFW Included"], 
                        label="Content Filter", 
                        value="SFW Only"
                    )
                    model_btn = gr.Button("Start Download", variant="primary")
                
                with gr.Column():
                    gr.Markdown("### Output Console")
                    model_output = gr.Textbox(label="Status", lines=20, interactive=False, autoscroll=True)
            
            # Toggle visibility based on mode selection
            mode.change(
                lambda x: [gr.update(visible=x == "single"), gr.update(visible=x == "bulk")],
                inputs=mode,
                outputs=[single_mode_group, bulk_mode_group]
            )
            
            model_btn.click(
                handle_model_download_new,
                inputs=[model_api_key, mode, model_input, file_input, nsfw_toggle],
                outputs=model_output
            )

        with gr.Tab("🖼️ Image Download"):
            with gr.Row():
                with gr.Column():
                    gr.Markdown("### Image Download Settings")
                    img_api_key = gr.Textbox(label="API Key", type="password")
                    img_url = gr.Textbox(label="Model Gallery URL", placeholder="https://civitai.com/models/...")
                    nsfw_mode = gr.Radio(["SFW Only", "NSFW Only"], label="Content Filter", value="SFW Only")
                    img_btn = gr.Button("Download Images", variant="primary")
        
                with gr.Column():
                    gr.Markdown("### Output Console")
                    img_output = gr.Textbox(label="Status", lines=20, interactive=False, autoscroll=True)
    
            # Connect the button click
            img_btn.click(
                handle_image_download,
                inputs=[img_api_key, img_url, nsfw_mode],
                outputs=img_output
            )

        with gr.Tab("📚 README"):
            gr.Markdown("""
            # 🚀 CivitAI Fetch
            
            **Developed by Voiid for Personal Use**
            
            ## 🔑 API Key Setup
            
            1. Log in to your CivitAI account
            2. Go to https://civitai.com/user/account
            3. Scroll down to "API Keys" section
            4. Generate a new API key
            5. Copy and paste it into the application
            
            ## 📦 Model Download Tab
            
            ### Single Download Mode:
            - Enter your API key
            - Paste a CivitAI model URL or model ID
            - Select content filter preference (SFW/NSFW)
            - Click "Start Download"
            
            ### Bulk Download Mode:
            - Create a text file with one model URL per line
            - Upload the text file
            - Select content filter preference
            - Click "Start Download"
            
            ## 🖼️ Image Download Tab
            
            - Enter your API key
            - Paste a CivitAI model gallery URL
            - Select content filter preference (SFW/NSFW)
            - Click "Download Images"
            
            ## 📋 Additional Information
            
            - Downloaded models are saved in the **CivitModels** folder
            - Downloaded images are saved in the **CivitImg** folder
            - Reports are generated in the **CivitData** folder
            - The application creates detailed Excel reports with model information
            
            ## 🔗 Links
            
            - [GitHub Repository](https://github.com/official-imvoiid)
            - CLI Version also available in the [CivitFetch repository](https://github.com/official-imvoiid/CivitFetch)
            
            ## ⚠️ Disclaimer
            
            This tool is for personal use only. The developer is not responsible for any actions taken with this tool or any downloaded content. Users are responsible for complying with CivitAI's terms of service and all applicable laws regarding the use and distribution of AI models and images.

            ### ⚠️ NSFW Disclaimer

            > **NOTE:** This script **may download NSFW content** if:
            >
            > - Your account or API key **has NSFW enabled**, **AND**
            > - The image/model being downloaded **doesn't include an explicit `NSFW` tag**
            
            Many models/images don’t label NSFW content properly, so there's a **chance of accidental downloads**.
            Building a script that 100% filters NSFW is extremely difficult due to inconsistent tagging.
            🔒 **If you want to avoid all NSFW content**, make sure to **disable NSFW in your account and API key settings**.
            Always **double-check downloaded files** before use.
                        
            ## 📄 License
            
            For personal use only. Not for commercial use or Redistribution. Project is Under MIT License
            """)

if __name__ == "__main__":
    app.launch()